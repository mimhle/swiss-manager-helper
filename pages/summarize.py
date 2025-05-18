from io import BytesIO
from operator import itemgetter
from typing import Literal

import dash
import dash_bootstrap_components as dbc
import pandas
import plotly.express as px
from dash import Output, Input, dcc
from dash.exceptions import PreventUpdate
from openpyxl import Workbook
from openpyxl.cell import Cell
from openpyxl.styles import Font, PatternFill

from components.table import table
from utils import parse_number, autofit_columns

dash.register_page(
    __name__,
    path='/summarize',
)

layout = dbc.Container([
    dbc.Accordion(
        [
            dbc.AccordionItem(
                dbc.Container([
                    table(
                        id="table_group_summarize",
                        columns=[{
                            "name": "Short name",
                            "id": "shortName",
                            "deletable": False,
                            "selectable": True,
                            "type": "text",
                            "editable": True,
                        }, {
                            "name": "Long name",
                            "id": "longName",
                            "deletable": False,
                            "selectable": True,
                            "type": "text",
                            "editable": True,
                        }],
                        data=[{"": 1}],
                    ),
                    dbc.Row([
                        dbc.Col(dbc.Button("Replace name long -> short", id="replace_lts", n_clicks=0, color="secondary", disabled=True, className="me-1 w-fit"), width="auto"),
                        dbc.Col(dbc.Button("Replace name short -> long", id="replace_stl", n_clicks=0, color="secondary", disabled=True, className="me-1 w-fit"), width="auto"),
                    ]),
                ], className="flex flex-col gap-2 p-0"), title="Convert team name",
            ),
        ],
        start_collapsed=True
    ),
    table(
        id="table_summarize",
        columns=[{
            "name": v,
            "id": k,
            "deletable": False,
            "selectable": True,
            "type": "text",
            "editable": True,
        } for k, v in {
            "rank": "Rank",
            "no": "No",
            "name": "Name",
            "team": "Team",
            "score": "Score",
            "tb1": "TB1",
            "tb2": "TB2",
            "tb3": "TB3",
            "tb4": "TB4",
            "tb5": "TB5",
        }.items()],
        data=[{"": 1}],
    ),
    dbc.Container([
        dbc.Label("Rank by:"),
        dbc.RadioItems(
            id="sort_by",
            options=[
                {"label": "Rank", "value": "rank"},
                {"label": "Score", "value": "score"},
            ],
            inline=True,
            value="rank",
        ),
    ]),
    dbc.Container([
        dcc.Graph(figure=px.bar(), id="graph_summarize", className="w-full inline-block"),
    ], className="w-full"),
    dbc.DropdownMenu([
        dbc.DropdownMenuItem(
            "Export to excel", id="export", n_clicks=0, className="me-1"
        ),
    ], label="Export", class_name="p-0", id="export_menu"),
    dash.dcc.Download(id="download-text"),
], className="flex flex-col gap-2 p-0")


@dash.callback(
    Output("table_summarize", "data", allow_duplicate=True),
    Input("table_summarize", "data"),
    prevent_initial_call=True,
)
def change_data(data):
    if not data:
        data = [{"": 1}]
        return data

    data = [
        row for row in data if row.get("rank") or row.get("no") or row.get("name") or row.get("team")
    ]

    data.append({"": 1})
    return data


@dash.callback(
    Output("table_group_summarize", "data", allow_duplicate=True),
    Output("replace_lts", "disabled"),
    Output("replace_stl", "disabled"),
    Input("table_group_summarize", "data"),
    prevent_initial_call=True,
)
def update_table_group_summarize(data):
    if not data:
        data = [{"": 1}]
        return data, True, True
    data = [
        row for row in data if row.get("longName") or row.get("shortName")
    ]
    data.append({"": 1})

    return data, False, False


@dash.callback(
    Output("table_summarize", "data", allow_duplicate=True),
    Input("replace_lts", "n_clicks"),
    Input("replace_stl", "n_clicks"),
    Input("table_group_summarize", "data"),
    Input("table_summarize", "data"),
    prevent_initial_call=True,
)
def replace_team_name(n_lts, n_stl, team_data, data):
    if dash.ctx.triggered_id == "table_summarize":
        raise PreventUpdate

    if dash.ctx.triggered_id == "replace_lts":
        team_data = {row["longName"]: row["shortName"] for row in team_data if row.get("longName") and row.get("shortName")}
        for row in data:
            if not row.get("team"):
                continue
            row["team"] = team_data.get(row["team"], row["team"])
    elif dash.ctx.triggered_id == "replace_stl":
        team_data = {row["shortName"]: row["longName"] for row in team_data if row.get("longName") and row.get("shortName")}
        for row in data:
            if not row.get("team"):
                continue
            row["team"] = team_data.get(row["team"], row["team"])
    else:
        raise PreventUpdate

    return data


def generate_summary(data, sort_by: Literal["rank", "score"] = "rank", top=3):
    result = {}
    for i, row in enumerate(data):
        if not row.get("rank") and not row.get("no") and not row.get("name") and not row.get("team"):
            continue

        if row.get("rank") == "":
            row["rank"] = data[i - 1]["rank"] if i > 0 else 1

        team = row["team"]
        if team not in result:
            result[team] = {"players": [], "rank": 0, "score": 0, "tb1": 0, "tb2": 0, "tb3": 0, "tb4": 0, "tb5": 0}
        result[team]["players"].append({k: parse_number(v) if k in ["rank", "score", "tb1", "tb2", "tb3", "tb4", "tb5"] else v for k, v in row.items()})

    for team in result:
        if sort_by == "rank":
            result[team]["players"] = sorted(result[team]["players"], key=itemgetter("rank"))[:top]
        elif sort_by == "score":
            result[team]["players"] = sorted(result[team]["players"], key=lambda x: (x["score"], 99999 - x["rank"]), reverse=True)[:top]

    for team in result:
        result[team]["rank"] = sum(player["rank"] for player in result[team]["players"])
        result[team]["score"] = sum(player["score"] for player in result[team]["players"])
        result[team]["tb1"] = sum(player["tb1"] for player in result[team]["players"])
        result[team]["tb2"] = sum(player["tb2"] for player in result[team]["players"])
        result[team]["tb3"] = sum(player["tb3"] for player in result[team]["players"])
        result[team]["tb4"] = sum(player["tb4"] for player in result[team]["players"])
        result[team]["tb5"] = sum(player["tb5"] for player in result[team]["players"])

    return result


@dash.callback(
    Output("graph_summarize", "figure"),
    Input("table_summarize", "data"),
    Input("sort_by", "value"),
    prevent_initial_call=True,
)
def update_graph(data, sort_by: Literal["rank", "score"] = "rank"):
    if not data:
        raise PreventUpdate

    summary_data = generate_summary(data, sort_by=sort_by, top=2)
    graph_data = pandas.DataFrame(columns=["team", "rank", "score", "tb1", "tb2", "tb3", "tb4", "tb5"])
    for team, values in sorted(summary_data.items(), key=lambda x: (len(x[1]["players"]), 99999 - x[1]["rank"], x[1]["score"], x[1]["tb1"], x[1]["tb2"], x[1]["tb3"], x[1]["tb4"], x[1]["tb5"]) if sort_by == "rank" else (len(x[1]["players"]), x[1]["score"], 99999 - x[1]["rank"], x[1]["tb1"], x[1]["tb2"], x[1]["tb3"], x[1]["tb4"], x[1]["tb5"]), reverse=True):
        graph_data = pandas.concat([graph_data, pandas.DataFrame([{
            "team": team,
            "rank": values["rank"],
            "score": values["score"],
            "tb1": values["tb1"],
            "tb2": values["tb2"],
            "tb3": values["tb3"],
            "tb4": values["tb4"],
            "tb5": values["tb5"],
        }])], ignore_index=True)

    fig = px.bar(
        graph_data,
        x="team",
        y=["rank", "score", "tb1", "tb2", "tb3", "tb4", "tb5"] if sort_by == "rank" else ["score", "rank", "tb1", "tb2", "tb3", "tb4", "tb5"],
        barmode="group",
        title="Summary of teams",
        labels={"value": "Value", "team": "Team"},
        color_discrete_sequence=px.colors.qualitative.Plotly,
    )

    return fig


@dash.callback(
    Output("download-text", "data"),
    Input("export", "n_clicks"),
    Input("table_summarize", "data"),
    Input("sort_by", "value"),
    prevent_initial_call=True,
)
def export_to_excel(n_clicks, data, sort_by: Literal["rank", "score"] = "rank", top: int = 2):
    if not data or dash.ctx.triggered_id != "export":
        raise PreventUpdate

    summary_data = generate_summary(data, sort_by=sort_by, top=top)

    def styled_cells(d):
        for c in d:
            c = Cell(ws, column=1, row=1, value=c)
            c.font = Font(bold=True)
            c.fill = PatternFill(start_color="BFBFBF", end_color="BFBFBF", fill_type="solid")
            yield c

    wb = Workbook()
    ws = wb.active
    if sort_by == "rank":
        ws.append(["Rank", "Team", "Total Rank", "Score", "TB1", "TB2", "TB3", "TB4", "TB5"])
    else:
        ws.append(["Rank", "Team", "Score", "Total Rank", "TB1", "TB2", "TB3", "TB4", "TB5"])
    i = 1
    for team, values in sorted(summary_data.items(), key=lambda x: (len(x[1]["players"]), 99999 - x[1]["rank"], x[1]["score"], x[1]["tb1"], x[1]["tb2"], x[1]["tb3"], x[1]["tb4"], x[1]["tb5"]) if sort_by == "rank" else (len(x[1]["players"]), x[1]["score"], 99999 - x[1]["rank"], x[1]["tb1"], x[1]["tb2"], x[1]["tb3"], x[1]["tb4"], x[1]["tb5"]), reverse=True):
        ws.append(styled_cells(map(lambda x: f'{x:g}' if isinstance(x, float) else str(x), [i, team, values["rank"], values["score"], values["tb1"], values["tb2"], values["tb3"], values["tb4"], values["tb5"]] if sort_by == "rank" else [i, team, values["score"], values["rank"], values["tb1"], values["tb2"], values["tb3"], values["tb4"], values["tb5"]])))
        for j, player in enumerate(values["players"]):
            ws.append(list(map(lambda x: f'{x:g}' if isinstance(x, float) else str(x), [j+1, player["name"], player["rank"], player["score"], player["tb1"], player["tb2"], player["tb3"], player["tb4"], player["tb5"]] if sort_by == "rank" else [j+1, player["name"], player["score"], player["rank"], player["tb1"], player["tb2"], player["tb3"], player["tb4"], player["tb5"]])))

        i += 1

    autofit_columns(ws)

    virtual_workbook = BytesIO()
    wb.save(virtual_workbook)
    return dcc.send_bytes(
        virtual_workbook.getvalue(),
        filename="summary.xlsx",
    )